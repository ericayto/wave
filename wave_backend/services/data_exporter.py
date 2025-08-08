"""
Data Exporter
Export trading data in multiple formats with comprehensive formatting options.
"""

import asyncio
from datetime import datetime, timedelta, date
from typing import Dict, List, Optional, Any, Tuple, Union
from dataclasses import dataclass, asdict
import logging
import pandas as pd
import numpy as np
from pathlib import Path
import json
import csv
import zipfile
from io import StringIO, BytesIO
import xml.etree.ElementTree as ET

# Excel export support
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from ..config.settings import get_settings
from ..services.event_bus import EventBus

logger = logging.getLogger(__name__)

@dataclass
class ExportConfiguration:
    """Configuration for data export."""
    export_format: str  # 'csv', 'json', 'excel', 'xml', 'parquet'
    include_headers: bool = True
    date_format: str = "%Y-%m-%d"
    decimal_places: int = 4
    include_metadata: bool = True
    compress_output: bool = False
    split_by_strategy: bool = False
    include_charts: bool = False  # For Excel exports
    custom_fields: Optional[List[str]] = None

@dataclass
class ExportResult:
    """Result of data export operation."""
    export_id: str
    export_type: str
    file_path: str
    file_size_bytes: int
    records_exported: int
    export_time: float
    generated_at: datetime
    metadata: Dict[str, Any]
    
    # Validation results
    data_integrity_check: bool = True
    export_errors: List[str] = None
    export_warnings: List[str] = None

@dataclass
class TaxReportData:
    """Structured data for tax reporting."""
    transaction_type: str  # 'buy', 'sell', 'dividend', 'interest'
    date: date
    symbol: str
    quantity: float
    price: float
    fees: float
    total_amount: float
    currency: str
    exchange: str
    strategy_id: str
    
    # Tax-specific fields
    cost_basis: Optional[float] = None
    gain_loss: Optional[float] = None
    holding_period: Optional[int] = None  # days
    wash_sale: bool = False

@dataclass
class IntegrationExport:
    """Data structured for external integrations."""
    integration_type: str  # 'personal_capital', 'mint', 'quickbooks', 'portfolio_tracker'
    data_format: str
    file_path: str
    mapping_config: Dict[str, str]
    validation_rules: List[str]

class DataExporter:
    """Export trading data in multiple formats."""
    
    def __init__(self, event_bus: EventBus):
        self.event_bus = event_bus
        self.settings = get_settings()
        
        # Export directory
        self.export_dir = Path("exports")
        self.export_dir.mkdir(exist_ok=True)
        
        # Create subdirectories
        (self.export_dir / "csv").mkdir(exist_ok=True)
        (self.export_dir / "json").mkdir(exist_ok=True)
        (self.export_dir / "excel").mkdir(exist_ok=True)
        (self.export_dir / "tax_reports").mkdir(exist_ok=True)
        (self.export_dir / "integrations").mkdir(exist_ok=True)
        
        # Export tracking
        self.export_history = []
        self.active_exports = {}
        
        # Supported formats and their configurations
        self.supported_formats = {
            'csv': {
                'extension': '.csv',
                'mime_type': 'text/csv',
                'supports_compression': True,
                'supports_multiple_sheets': False
            },
            'json': {
                'extension': '.json',
                'mime_type': 'application/json',
                'supports_compression': True,
                'supports_multiple_sheets': False
            },
            'excel': {
                'extension': '.xlsx',
                'mime_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'supports_compression': False,
                'supports_multiple_sheets': True,
                'supports_charts': True
            },
            'xml': {
                'extension': '.xml',
                'mime_type': 'application/xml',
                'supports_compression': True,
                'supports_multiple_sheets': False
            },
            'parquet': {
                'extension': '.parquet',
                'mime_type': 'application/octet-stream',
                'supports_compression': True,
                'supports_multiple_sheets': False
            }
        }
        
        # Data types and their export configurations
        self.data_types = {
            'trades': {
                'table_name': 'trades',
                'required_fields': ['timestamp', 'symbol', 'side', 'quantity', 'price'],
                'optional_fields': ['fees', 'strategy_id', 'order_id', 'exchange']
            },
            'positions': {
                'table_name': 'positions',
                'required_fields': ['timestamp', 'symbol', 'quantity', 'market_value'],
                'optional_fields': ['unrealized_pnl', 'cost_basis', 'strategy_id']
            },
            'portfolio_snapshots': {
                'table_name': 'portfolio_snapshots',
                'required_fields': ['timestamp', 'total_value', 'cash_balance'],
                'optional_fields': ['realized_pnl', 'unrealized_pnl', 'margin_used']
            },
            'performance_metrics': {
                'table_name': 'performance_metrics',
                'required_fields': ['date', 'strategy_id', 'return', 'cumulative_return'],
                'optional_fields': ['sharpe_ratio', 'max_drawdown', 'volatility']
            },
            'orders': {
                'table_name': 'orders',
                'required_fields': ['timestamp', 'order_id', 'symbol', 'side', 'quantity'],
                'optional_fields': ['price', 'order_type', 'status', 'strategy_id']
            },
            'risk_metrics': {
                'table_name': 'risk_metrics',
                'required_fields': ['date', 'portfolio_value', 'var_95'],
                'optional_fields': ['var_99', 'expected_shortfall', 'beta', 'correlation']
            }
        }
        
    async def export_to_csv(self, 
                          data_type: str, 
                          date_range: Tuple[date, date],
                          config: Optional[ExportConfiguration] = None) -> ExportResult:
        """Export various data types to CSV format."""
        
        start_time = datetime.utcnow()
        
        if config is None:
            config = ExportConfiguration(export_format='csv')
        
        logger.info(f"Exporting {data_type} to CSV for date range {date_range[0]} to {date_range[1]}")
        
        # Fetch data
        data = await self._fetch_data(data_type, date_range)
        
        if not data:
            logger.warning(f"No data found for {data_type} in specified date range")
            return self._create_empty_export_result('csv', data_type)
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # Apply formatting
        df = await self._format_dataframe(df, config)
        
        # Generate filename
        filename = self._generate_filename(data_type, 'csv', date_range)
        file_path = self.export_dir / "csv" / filename
        
        # Export to CSV
        df.to_csv(
            file_path,
            index=False,
            date_format=config.date_format,
            float_format=f'%.{config.decimal_places}f'
        )
        
        # Compress if requested
        if config.compress_output:
            file_path = await self._compress_file(file_path)
        
        # Create result
        result = ExportResult(
            export_id=self._generate_export_id(),
            export_type=f'{data_type}_csv',
            file_path=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            records_exported=len(df),
            export_time=(datetime.utcnow() - start_time).total_seconds(),
            generated_at=datetime.utcnow(),
            metadata={
                'data_type': data_type,
                'date_range': date_range,
                'format': 'csv',
                'columns': list(df.columns),
                'compressed': config.compress_output
            }
        )
        
        # Validate export
        await self._validate_export(result, df)
        
        logger.info(f"CSV export completed: {file_path}")
        
        return result
    
    async def export_to_json(self, 
                           data_type: str,
                           date_range: Optional[Tuple[date, date]] = None,
                           config: Optional[ExportConfiguration] = None) -> ExportResult:
        """Export data to JSON format with hierarchical structure."""
        
        start_time = datetime.utcnow()
        
        if config is None:
            config = ExportConfiguration(export_format='json')
        
        logger.info(f"Exporting {data_type} to JSON")
        
        # Fetch data
        if date_range:
            data = await self._fetch_data(data_type, date_range)
        else:
            data = await self._fetch_all_data(data_type)
        
        if not data:
            return self._create_empty_export_result('json', data_type)
        
        # Structure data for JSON export
        json_data = await self._structure_json_data(data, data_type, config)
        
        # Generate filename
        filename = self._generate_filename(data_type, 'json', date_range)
        file_path = self.export_dir / "json" / filename
        
        # Export to JSON
        with open(file_path, 'w') as f:
            json.dump(json_data, f, indent=2, default=self._json_serializer)
        
        # Compress if requested
        if config.compress_output:
            file_path = await self._compress_file(file_path)
        
        # Create result
        result = ExportResult(
            export_id=self._generate_export_id(),
            export_type=f'{data_type}_json',
            file_path=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            records_exported=len(data),
            export_time=(datetime.utcnow() - start_time).total_seconds(),
            generated_at=datetime.utcnow(),
            metadata={
                'data_type': data_type,
                'date_range': date_range,
                'format': 'json',
                'structure': 'hierarchical',
                'compressed': config.compress_output
            }
        )
        
        logger.info(f"JSON export completed: {file_path}")
        
        return result
    
    async def export_to_excel(self,
                            data_types: List[str],
                            date_range: Tuple[date, date],
                            config: Optional[ExportConfiguration] = None) -> ExportResult:
        """Export multiple data types to Excel workbook with charts."""
        
        start_time = datetime.utcnow()
        
        if config is None:
            config = ExportConfiguration(export_format='excel', include_charts=True)
        
        logger.info(f"Exporting {data_types} to Excel workbook")
        
        # Generate filename
        filename = self._generate_filename('_'.join(data_types), 'xlsx', date_range)
        file_path = self.export_dir / "excel" / filename
        
        # Create workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet
        
        total_records = 0
        
        # Add data sheets
        for data_type in data_types:
            # Fetch data
            data = await self._fetch_data(data_type, date_range)
            
            if not data:
                continue
            
            df = pd.DataFrame(data)
            df = await self._format_dataframe(df, config)
            
            # Create worksheet
            worksheet = workbook.create_sheet(title=data_type.replace('_', ' ').title())
            
            # Add data
            for row in dataframe_to_rows(df, index=False, header=True):
                worksheet.append(row)
            
            # Format worksheet
            await self._format_excel_worksheet(worksheet, df, config)
            
            # Add charts if requested
            if config.include_charts and data_type in ['trades', 'performance_metrics', 'portfolio_snapshots']:
                await self._add_excel_charts(worksheet, df, data_type)
            
            total_records += len(df)
        
        # Add summary sheet
        await self._create_excel_summary_sheet(workbook, data_types, date_range)
        
        # Save workbook
        workbook.save(file_path)
        
        # Create result
        result = ExportResult(
            export_id=self._generate_export_id(),
            export_type='excel_workbook',
            file_path=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            records_exported=total_records,
            export_time=(datetime.utcnow() - start_time).total_seconds(),
            generated_at=datetime.utcnow(),
            metadata={
                'data_types': data_types,
                'date_range': date_range,
                'format': 'excel',
                'sheets': len(workbook.worksheets),
                'includes_charts': config.include_charts
            }
        )
        
        logger.info(f"Excel export completed: {file_path}")
        
        return result
    
    async def export_for_tax_reporting(self, year: int) -> ExportResult:
        """Export data formatted specifically for tax reporting."""
        
        start_time = datetime.utcnow()
        
        logger.info(f"Generating tax report for year {year}")
        
        # Date range for the tax year
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        
        # Fetch relevant data
        trades_data = await self._fetch_data('trades', (start_date, end_date))
        
        # Process for tax reporting
        tax_data = []
        
        for trade in trades_data:
            tax_record = await self._process_trade_for_tax(trade, year)
            if tax_record:
                tax_data.append(tax_record)
        
        # Generate tax report files
        results = []
        
        # 1. Detailed transaction report
        detailed_result = await self._create_detailed_tax_report(tax_data, year)
        results.append(detailed_result)
        
        # 2. Form 8949 format (capital gains/losses)
        form8949_result = await self._create_form8949_report(tax_data, year)
        results.append(form8949_result)
        
        # 3. Summary report
        summary_result = await self._create_tax_summary_report(tax_data, year)
        results.append(summary_result)
        
        # Create combined result
        result = ExportResult(
            export_id=self._generate_export_id(),
            export_type='tax_reporting',
            file_path=str(self.export_dir / "tax_reports" / f"tax_package_{year}.zip"),
            file_size_bytes=sum(r.file_size_bytes for r in results),
            records_exported=len(tax_data),
            export_time=(datetime.utcnow() - start_time).total_seconds(),
            generated_at=datetime.utcnow(),
            metadata={
                'tax_year': year,
                'reports_included': ['detailed', 'form8949', 'summary'],
                'total_transactions': len(tax_data),
                'realized_gains': sum(r.gain_loss for r in tax_data if r.gain_loss and r.gain_loss > 0),
                'realized_losses': sum(r.gain_loss for r in tax_data if r.gain_loss and r.gain_loss < 0)
            }
        )
        
        # Create zip package
        await self._create_tax_report_package(results, result.file_path)
        
        logger.info(f"Tax reporting export completed: {result.file_path}")
        
        return result
    
    async def integration_with_portfolio_trackers(self, 
                                                integration_type: str,
                                                date_range: Optional[Tuple[date, date]] = None) -> ExportResult:
        """Export data for integration with external portfolio tracking systems."""
        
        start_time = datetime.utcnow()
        
        logger.info(f"Creating export for {integration_type} integration")
        
        # Integration configurations
        integration_configs = {
            'personal_capital': {
                'format': 'csv',
                'required_fields': ['Date', 'Account', 'Symbol', 'Description', 'Quantity', 'Price', 'Amount'],
                'date_format': '%m/%d/%Y'
            },
            'mint': {
                'format': 'csv',
                'required_fields': ['Date', 'Description', 'Original Description', 'Amount', 'Transaction Type', 'Category', 'Account Name'],
                'date_format': '%m/%d/%Y'
            },
            'quickbooks': {
                'format': 'csv',
                'required_fields': ['Date', 'Account', 'Debit', 'Credit', 'Description', 'Name'],
                'date_format': '%m/%d/%Y'
            },
            'portfolio_tracker': {
                'format': 'json',
                'structure': 'hierarchical',
                'includes_performance': True
            }
        }
        
        if integration_type not in integration_configs:
            raise ValueError(f"Unsupported integration type: {integration_type}")
        
        config = integration_configs[integration_type]
        
        # Fetch and transform data
        if date_range:
            raw_data = await self._fetch_data('trades', date_range)
        else:
            raw_data = await self._fetch_all_data('trades')
        
        # Transform data for integration
        transformed_data = await self._transform_for_integration(raw_data, integration_type, config)
        
        # Generate filename and export
        filename = f"{integration_type}_export_{datetime.utcnow().strftime('%Y%m%d')}"
        
        if config['format'] == 'csv':
            file_path = self.export_dir / "integrations" / f"{filename}.csv"
            df = pd.DataFrame(transformed_data)
            df.to_csv(file_path, index=False, date_format=config['date_format'])
        else:
            file_path = self.export_dir / "integrations" / f"{filename}.json"
            with open(file_path, 'w') as f:
                json.dump(transformed_data, f, indent=2, default=self._json_serializer)
        
        # Create result
        result = ExportResult(
            export_id=self._generate_export_id(),
            export_type=f'{integration_type}_integration',
            file_path=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            records_exported=len(transformed_data),
            export_time=(datetime.utcnow() - start_time).total_seconds(),
            generated_at=datetime.utcnow(),
            metadata={
                'integration_type': integration_type,
                'format': config['format'],
                'date_range': date_range,
                'mapping_applied': True
            }
        )
        
        logger.info(f"Integration export completed: {file_path}")
        
        return result
    
    async def export_custom_query(self,
                                query: str,
                                export_format: str = 'csv',
                                config: Optional[ExportConfiguration] = None) -> ExportResult:
        """Export results of custom data query."""
        
        start_time = datetime.utcnow()
        
        logger.info(f"Executing custom query export to {export_format}")
        
        # Execute custom query (mock implementation)
        data = await self._execute_custom_query(query)
        
        if not data:
            return self._create_empty_export_result(export_format, 'custom_query')
        
        # Export based on format
        if export_format == 'csv':
            result = await self._export_data_to_csv(data, 'custom_query', config)
        elif export_format == 'json':
            result = await self._export_data_to_json(data, 'custom_query', config)
        elif export_format == 'excel':
            result = await self._export_data_to_excel(data, 'custom_query', config)
        else:
            raise ValueError(f"Unsupported export format: {export_format}")
        
        result.export_time = (datetime.utcnow() - start_time).total_seconds()
        result.metadata['custom_query'] = query
        
        return result
    
    # Private helper methods
    
    async def _fetch_data(self, data_type: str, date_range: Tuple[date, date]) -> List[Dict]:
        """Fetch data for specified type and date range."""
        
        # Mock data generation based on data type
        start_date, end_date = date_range
        days = (end_date - start_date).days + 1
        
        if data_type == 'trades':
            return [
                {
                    'timestamp': start_date + timedelta(days=i, hours=np.random.randint(0, 24)),
                    'symbol': np.random.choice(['BTC-USD', 'ETH-USD', 'ADA-USD']),
                    'side': np.random.choice(['buy', 'sell']),
                    'quantity': np.random.uniform(0.1, 5.0),
                    'price': np.random.uniform(20000, 50000),
                    'fees': np.random.uniform(1, 50),
                    'strategy_id': f'strategy_{np.random.randint(1, 4)}',
                    'order_id': f'order_{i}_{np.random.randint(1000, 9999)}',
                    'exchange': 'coinbase_pro'
                }
                for i in range(min(days * 5, 1000))  # 5 trades per day on average
            ]
        
        elif data_type == 'positions':
            return [
                {
                    'timestamp': start_date + timedelta(days=i),
                    'symbol': symbol,
                    'quantity': np.random.uniform(0.5, 10.0),
                    'market_value': np.random.uniform(1000, 50000),
                    'unrealized_pnl': np.random.uniform(-1000, 2000),
                    'cost_basis': np.random.uniform(15000, 45000),
                    'strategy_id': f'strategy_{np.random.randint(1, 4)}'
                }
                for i in range(days)
                for symbol in ['BTC-USD', 'ETH-USD']
            ]
        
        elif data_type == 'performance_metrics':
            returns = np.random.normal(0.001, 0.02, days)
            cumulative_returns = np.cumprod(1 + returns) - 1
            
            return [
                {
                    'date': start_date + timedelta(days=i),
                    'strategy_id': f'strategy_{j}',
                    'return': returns[i] + np.random.normal(0, 0.005),
                    'cumulative_return': cumulative_returns[i],
                    'sharpe_ratio': np.random.uniform(0.5, 2.5),
                    'max_drawdown': np.random.uniform(0.02, 0.15),
                    'volatility': np.random.uniform(0.15, 0.30)
                }
                for i in range(days)
                for j in range(1, 4)
            ]
        
        else:
            # Generic data structure
            return [
                {
                    'timestamp': start_date + timedelta(days=i),
                    'value': np.random.uniform(100, 1000),
                    'category': f'category_{np.random.randint(1, 5)}'
                }
                for i in range(days)
            ]
    
    async def _fetch_all_data(self, data_type: str) -> List[Dict]:
        """Fetch all available data for specified type."""
        # Use a default range of last 90 days
        end_date = date.today()
        start_date = end_date - timedelta(days=90)
        return await self._fetch_data(data_type, (start_date, end_date))
    
    async def _format_dataframe(self, df: pd.DataFrame, config: ExportConfiguration) -> pd.DataFrame:
        """Format DataFrame according to export configuration."""
        
        # Round numeric columns
        numeric_columns = df.select_dtypes(include=[np.number]).columns
        df[numeric_columns] = df[numeric_columns].round(config.decimal_places)
        
        # Format datetime columns
        datetime_columns = df.select_dtypes(include=['datetime64']).columns
        for col in datetime_columns:
            df[col] = df[col].dt.strftime(config.date_format)
        
        # Include only custom fields if specified
        if config.custom_fields:
            available_fields = [f for f in config.custom_fields if f in df.columns]
            df = df[available_fields]
        
        return df
    
    def _generate_filename(self, data_type: str, extension: str, date_range: Optional[Tuple[date, date]] = None) -> str:
        """Generate filename for export."""
        
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        
        if date_range:
            start_str = date_range[0].strftime('%Y%m%d')
            end_str = date_range[1].strftime('%Y%m%d')
            return f"{data_type}_{start_str}_to_{end_str}_{timestamp}.{extension}"
        else:
            return f"{data_type}_{timestamp}.{extension}"
    
    def _generate_export_id(self) -> str:
        """Generate unique export ID."""
        return f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{np.random.randint(1000, 9999)}"
    
    async def _compress_file(self, file_path: Path) -> Path:
        """Compress file using ZIP."""
        
        zip_path = file_path.with_suffix(file_path.suffix + '.zip')
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(file_path, file_path.name)
        
        # Remove original file
        file_path.unlink()
        
        return zip_path
    
    def _json_serializer(self, obj):
        """JSON serializer for datetime objects."""
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        return str(obj)
    
    async def _structure_json_data(self, data: List[Dict], data_type: str, config: ExportConfiguration) -> Dict:
        """Structure data for hierarchical JSON export."""
        
        structured = {
            'metadata': {
                'data_type': data_type,
                'generated_at': datetime.utcnow().isoformat(),
                'record_count': len(data),
                'export_config': asdict(config)
            },
            'data': data
        }
        
        # Group by strategy if requested
        if config.split_by_strategy and 'strategy_id' in (data[0] if data else {}):
            grouped_data = {}
            for record in data:
                strategy_id = record.get('strategy_id', 'unknown')
                if strategy_id not in grouped_data:
                    grouped_data[strategy_id] = []
                grouped_data[strategy_id].append(record)
            structured['data'] = grouped_data
        
        return structured
    
    def _create_empty_export_result(self, format_type: str, data_type: str) -> ExportResult:
        """Create empty export result for cases with no data."""
        
        return ExportResult(
            export_id=self._generate_export_id(),
            export_type=f'{data_type}_{format_type}',
            file_path='',
            file_size_bytes=0,
            records_exported=0,
            export_time=0.0,
            generated_at=datetime.utcnow(),
            metadata={'status': 'no_data_found'},
            export_warnings=['No data found for specified criteria']
        )
    
    async def _validate_export(self, result: ExportResult, df: pd.DataFrame):
        """Validate export integrity."""
        
        errors = []
        warnings = []
        
        # Check for data consistency
        if len(df) != result.records_exported:
            errors.append(f"Record count mismatch: {len(df)} vs {result.records_exported}")
        
        # Check for missing critical data
        if df.isnull().sum().sum() > len(df) * 0.1:  # More than 10% null values
            warnings.append("High percentage of null values detected")
        
        result.export_errors = errors
        result.export_warnings = warnings
        result.data_integrity_check = len(errors) == 0
    
    # Additional methods for specific export types
    
    async def _format_excel_worksheet(self, worksheet, df: pd.DataFrame, config: ExportConfiguration):
        """Format Excel worksheet with styling."""
        
        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for cell in worksheet[1]:  # First row (header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    async def _add_excel_charts(self, worksheet, df: pd.DataFrame, data_type: str):
        """Add charts to Excel worksheet."""
        
        if data_type == 'performance_metrics' and 'return' in df.columns:
            # Create line chart for returns
            chart = LineChart()
            chart.title = "Performance Returns"
            chart.style = 13
            chart.x_axis.title = 'Date'
            chart.y_axis.title = 'Return'
            
            # Add data to chart (simplified)
            data = Reference(worksheet, min_col=3, min_row=1, max_row=min(100, len(df) + 1), max_col=3)
            chart.add_data(data, titles_from_data=True)
            
            worksheet.add_chart(chart, "F2")
    
    async def _create_excel_summary_sheet(self, workbook, data_types: List[str], date_range: Tuple[date, date]):
        """Create summary sheet for Excel workbook."""
        
        summary_sheet = workbook.create_sheet(title="Summary", index=0)
        
        # Add summary information
        summary_data = [
            ["Export Summary", ""],
            ["Generated At", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")],
            ["Date Range", f"{date_range[0]} to {date_range[1]}"],
            ["Data Types", ", ".join(data_types)],
            ["", ""],
            ["Sheet", "Records"],
        ]
        
        # Add sheet information
        for sheet_name in workbook.sheetnames[1:]:  # Skip summary sheet
            sheet = workbook[sheet_name]
            record_count = sheet.max_row - 1  # Subtract header row
            summary_data.append([sheet_name, record_count])
        
        # Write summary data
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # Title row
                    cell.font = Font(bold=True, size=14)
                elif row_idx == 6:  # Header row for sheet info
                    cell.font = Font(bold=True)
    
    # Tax reporting specific methods
    
    async def _process_trade_for_tax(self, trade: Dict, tax_year: int) -> Optional[TaxReportData]:
        """Process individual trade for tax reporting."""
        
        trade_date = trade['timestamp']
        if isinstance(trade_date, str):
            trade_date = datetime.strptime(trade_date, '%Y-%m-%d').date()
        elif isinstance(trade_date, datetime):
            trade_date = trade_date.date()
        
        # Only include trades from the tax year
        if trade_date.year != tax_year:
            return None
        
        # Calculate tax-specific fields
        cost_basis = trade['quantity'] * trade['price'] + trade.get('fees', 0)
        
        return TaxReportData(
            transaction_type='buy' if trade['side'] == 'buy' else 'sell',
            date=trade_date,
            symbol=trade['symbol'],
            quantity=trade['quantity'],
            price=trade['price'],
            fees=trade.get('fees', 0),
            total_amount=cost_basis,
            currency='USD',
            exchange=trade.get('exchange', 'unknown'),
            strategy_id=trade.get('strategy_id', 'unknown'),
            cost_basis=cost_basis if trade['side'] == 'buy' else None,
            gain_loss=0 if trade['side'] == 'buy' else np.random.uniform(-100, 200),  # Mock calculation
            holding_period=np.random.randint(1, 365) if trade['side'] == 'sell' else None,
            wash_sale=False  # Would need complex calculation
        )
    
    async def _create_detailed_tax_report(self, tax_data: List[TaxReportData], year: int) -> ExportResult:
        """Create detailed tax transaction report."""
        
        df = pd.DataFrame([asdict(td) for td in tax_data])
        
        filename = f"detailed_tax_report_{year}.csv"
        file_path = self.export_dir / "tax_reports" / filename
        
        df.to_csv(file_path, index=False)
        
        return ExportResult(
            export_id=self._generate_export_id(),
            export_type='detailed_tax_report',
            file_path=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            records_exported=len(df),
            export_time=0.0,
            generated_at=datetime.utcnow(),
            metadata={'tax_year': year, 'report_type': 'detailed'}
        )
    
    async def _create_form8949_report(self, tax_data: List[TaxReportData], year: int) -> ExportResult:
        """Create Form 8949 compatible report."""
        
        # Filter only sales transactions
        sales_data = [td for td in tax_data if td.transaction_type == 'sell']
        
        form8949_data = []
        for sale in sales_data:
            form8949_data.append({
                'Description': f"{sale.quantity} {sale.symbol}",
                'Date Acquired': '',  # Would need to look up
                'Date Sold': sale.date.strftime('%m/%d/%Y'),
                'Proceeds': sale.total_amount,
                'Cost Basis': sale.cost_basis or 0,
                'Gain/Loss': sale.gain_loss or 0,
                'Short/Long Term': 'Short' if (sale.holding_period or 0) <= 365 else 'Long'
            })
        
        df = pd.DataFrame(form8949_data)
        
        filename = f"form8949_report_{year}.csv"
        file_path = self.export_dir / "tax_reports" / filename
        
        df.to_csv(file_path, index=False)
        
        return ExportResult(
            export_id=self._generate_export_id(),
            export_type='form8949_report',
            file_path=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            records_exported=len(df),
            export_time=0.0,
            generated_at=datetime.utcnow(),
            metadata={'tax_year': year, 'report_type': 'form8949'}
        )
    
    async def _create_tax_summary_report(self, tax_data: List[TaxReportData], year: int) -> ExportResult:
        """Create tax summary report."""
        
        summary = {
            'tax_year': year,
            'total_transactions': len(tax_data),
            'total_buys': len([td for td in tax_data if td.transaction_type == 'buy']),
            'total_sells': len([td for td in tax_data if td.transaction_type == 'sell']),
            'total_realized_gains': sum(td.gain_loss for td in tax_data if td.gain_loss and td.gain_loss > 0),
            'total_realized_losses': sum(td.gain_loss for td in tax_data if td.gain_loss and td.gain_loss < 0),
            'net_gain_loss': sum(td.gain_loss for td in tax_data if td.gain_loss),
            'total_fees_paid': sum(td.fees for td in tax_data)
        }
        
        filename = f"tax_summary_{year}.json"
        file_path = self.export_dir / "tax_reports" / filename
        
        with open(file_path, 'w') as f:
            json.dump(summary, f, indent=2, default=self._json_serializer)
        
        return ExportResult(
            export_id=self._generate_export_id(),
            export_type='tax_summary_report',
            file_path=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            records_exported=1,  # One summary record
            export_time=0.0,
            generated_at=datetime.utcnow(),
            metadata={'tax_year': year, 'report_type': 'summary'}
        )
    
    async def _create_tax_report_package(self, reports: List[ExportResult], zip_path: str):
        """Create ZIP package of all tax reports."""
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for report in reports:
                if Path(report.file_path).exists():
                    zipf.write(report.file_path, Path(report.file_path).name)
    
    # Integration specific methods
    
    async def _transform_for_integration(self, data: List[Dict], integration_type: str, config: Dict) -> List[Dict]:
        """Transform data for specific integration format."""
        
        if integration_type == 'personal_capital':
            return [
                {
                    'Date': trade['timestamp'].strftime(config['date_format']) if isinstance(trade['timestamp'], datetime) else trade['timestamp'],
                    'Account': 'Trading Account',
                    'Symbol': trade['symbol'],
                    'Description': f"{trade['side'].title()} {trade['quantity']} {trade['symbol']}",
                    'Quantity': trade['quantity'] if trade['side'] == 'buy' else -trade['quantity'],
                    'Price': trade['price'],
                    'Amount': trade['quantity'] * trade['price'] * (1 if trade['side'] == 'buy' else -1)
                }
                for trade in data
            ]
        
        elif integration_type == 'mint':
            return [
                {
                    'Date': trade['timestamp'].strftime(config['date_format']) if isinstance(trade['timestamp'], datetime) else trade['timestamp'],
                    'Description': f"{trade['side'].title()} {trade['quantity']} {trade['symbol']}",
                    'Original Description': f"{trade['side'].title()} {trade['quantity']} {trade['symbol']}",
                    'Amount': trade['quantity'] * trade['price'] * (-1 if trade['side'] == 'buy' else 1),
                    'Transaction Type': 'debit' if trade['side'] == 'buy' else 'credit',
                    'Category': 'Investments',
                    'Account Name': 'Trading Account'
                }
                for trade in data
            ]
        
        else:
            # Generic transformation
            return data
    
    # Query execution (mock implementation)
    
    async def _execute_custom_query(self, query: str) -> List[Dict]:
        """Execute custom data query."""
        
        # Mock implementation - would execute actual query
        logger.info(f"Executing custom query: {query}")
        
        # Return mock data
        return [
            {
                'field1': f'value_{i}',
                'field2': np.random.uniform(100, 1000),
                'field3': datetime.utcnow() - timedelta(days=i)
            }
            for i in range(100)
        ]
    
    async def _export_data_to_csv(self, data: List[Dict], data_type: str, config: Optional[ExportConfiguration]) -> ExportResult:
        """Export generic data to CSV."""
        
        df = pd.DataFrame(data)
        filename = self._generate_filename(data_type, 'csv')
        file_path = self.export_dir / "csv" / filename
        
        df.to_csv(file_path, index=False)
        
        return ExportResult(
            export_id=self._generate_export_id(),
            export_type=f'{data_type}_csv',
            file_path=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            records_exported=len(df),
            export_time=0.0,
            generated_at=datetime.utcnow(),
            metadata={'data_type': data_type, 'format': 'csv'}
        )
    
    async def _export_data_to_json(self, data: List[Dict], data_type: str, config: Optional[ExportConfiguration]) -> ExportResult:
        """Export generic data to JSON."""
        
        filename = self._generate_filename(data_type, 'json')
        file_path = self.export_dir / "json" / filename
        
        with open(file_path, 'w') as f:
            json.dump(data, f, indent=2, default=self._json_serializer)
        
        return ExportResult(
            export_id=self._generate_export_id(),
            export_type=f'{data_type}_json',
            file_path=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            records_exported=len(data),
            export_time=0.0,
            generated_at=datetime.utcnow(),
            metadata={'data_type': data_type, 'format': 'json'}
        )
    
    async def _export_data_to_excel(self, data: List[Dict], data_type: str, config: Optional[ExportConfiguration]) -> ExportResult:
        """Export generic data to Excel."""
        
        df = pd.DataFrame(data)
        filename = self._generate_filename(data_type, 'xlsx')
        file_path = self.export_dir / "excel" / filename
        
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=data_type)
        
        return ExportResult(
            export_id=self._generate_export_id(),
            export_type=f'{data_type}_excel',
            file_path=str(file_path),
            file_size_bytes=file_path.stat().st_size,
            records_exported=len(df),
            export_time=0.0,
            generated_at=datetime.utcnow(),
            metadata={'data_type': data_type, 'format': 'excel'}
        )
    
    def get_export_status(self, export_id: str) -> Dict[str, Any]:
        """Get status of export operation."""
        
        return {
            'export_id': export_id,
            'status': 'completed',
            'progress': 1.0,
            'estimated_completion': None,
            'current_operation': 'export_complete'
        }
    
    def list_export_history(self, limit: int = 50) -> List[Dict]:
        """Get history of export operations."""
        
        return [
            {
                'export_id': f'export_{i}',
                'export_type': np.random.choice(['trades_csv', 'performance_json', 'tax_report']),
                'generated_at': datetime.utcnow() - timedelta(days=i),
                'file_size_bytes': np.random.randint(1024, 1024*1024),
                'records_exported': np.random.randint(100, 10000),
                'status': 'completed'
            }
            for i in range(min(limit, 50))
        ]